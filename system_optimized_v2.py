"""
Sistema de Asistencia Optimizado con PostgreSQL
Versión actualizada que usa la configuración de .env
"""
from flask import Flask, render_template, jsonify, request, send_file
from flask_socketio import SocketIO, emit
from flask_cors import CORS
import requests
from requests.auth import HTTPDigestAuth
import json
import threading
from datetime import datetime, timedelta, time as dt_time
import time as time_module
import os
from dotenv import load_dotenv
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️ openpyxl no disponible - exportación Excel deshabilitada")

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    print("⚠️ reportlab no disponible - exportación PDF deshabilitada")

import io

# Cargar variables de entorno
load_dotenv()

app = Flask(__name__, static_folder='static')
app.secret_key = os.getenv('FLASK_SECRET_KEY', 'hikvision_attendance_2024')
CORS(app, origins=["http://localhost:3001", "http://localhost:3000"])
socketio = SocketIO(app, cors_allowed_origins="*")

class OptimizedAttendanceSystem:
    def __init__(self):
        # Configuración desde variables de entorno
        self.device_ip = os.getenv('DEVICE_IP', '172.10.1.62')
        self.username = os.getenv('DEVICE_USER', 'admin')
        self.password = os.getenv('DEVICE_PASS', 'PC2024*+')
        self.database_url = os.getenv('DATABASE_URL')
        
        # Configurar conexión HTTP
        self.session = requests.Session()
        self.session.auth = HTTPDigestAuth(self.username, self.password)
        self.session.timeout = 10
        self.base_url = f"http://{self.device_ip}/ISAPI"
        
        # Estado del sistema
        self.monitoring = False
        self.connected = False
        self.last_event_time = time_module.time()
        
        # Cache para optimización
        self.employees_cache = {}
        self.cache_timestamp = 0
        self.cache_duration = 300  # 5 minutos
        
        # Configurar base de datos
        self.setup_database()
        
    def setup_database(self):
        """Configurar conexión a base de datos"""
        if self.database_url and self.database_url.startswith('postgresql'):
            # Usar PostgreSQL
            import psycopg2
            self.db_type = 'postgresql'
            self.get_connection = lambda: psycopg2.connect(self.database_url)
            print("Configurado para PostgreSQL")
        else:
            # Fallback a SQLite
            import sqlite3
            self.db_type = 'sqlite'
            self.db_path = 'attendance.db'
            self.get_connection = lambda: sqlite3.connect(self.db_path)
            print("Usando SQLite como fallback")
        
        self.init_database()
        
    def init_database(self):
        """Inicializar base de datos"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            if self.db_type == 'postgresql':
                # Crear tablas PostgreSQL
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS employees (
                        id SERIAL PRIMARY KEY,
                        employee_id TEXT UNIQUE NOT NULL,
                        name TEXT NOT NULL,
                        department TEXT DEFAULT 'General',
                        schedule TEXT DEFAULT 'estandar',
                        phone TEXT DEFAULT '',
                        email TEXT DEFAULT '',
                        active BOOLEAN DEFAULT true,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        synced_to_device BOOLEAN DEFAULT false
                    )
                ''')
                
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS employee_schedules (
                        id SERIAL PRIMARY KEY,
                        employee_id TEXT NOT NULL,
                        schedule_type TEXT NOT NULL,
                        shift_type TEXT DEFAULT NULL,
                        start_time TIME NOT NULL,
                        end_time TIME NOT NULL,
                        days_of_week TEXT NOT NULL,
                        active_from DATE DEFAULT CURRENT_DATE,
                        active_until DATE DEFAULT NULL,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                ''')
                
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS break_types (
                        id SERIAL PRIMARY KEY,
                        name VARCHAR(50) NOT NULL,
                        display_name VARCHAR(100) NOT NULL,
                        duration_minutes INTEGER NOT NULL,
                        mandatory BOOLEAN DEFAULT true,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                ''')
                
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS department_schedules (
                        id SERIAL PRIMARY KEY,
                        department VARCHAR(100) NOT NULL,
                        shift_type VARCHAR(50),
                        work_start TIME NOT NULL,
                        work_end TIME NOT NULL,
                        break_start TIME NOT NULL,
                        break_end TIME NOT NULL,
                        has_lunch BOOLEAN DEFAULT FALSE,
                        lunch_options TEXT[],
                        friday_end TIME,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                ''')
                
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS attendance_records (
                        id SERIAL PRIMARY KEY,
                        employee_id TEXT NOT NULL,
                        event_type TEXT NOT NULL,
                        timestamp TIMESTAMP NOT NULL,
                        reader_no INTEGER DEFAULT 1,
                        verify_method TEXT DEFAULT 'huella',
                        status TEXT DEFAULT 'autorizado',
                        break_type VARCHAR(50),
                        is_break_record BOOLEAN DEFAULT FALSE,
                        break_duration_minutes INTEGER
                    )
                ''')
                
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS daily_summaries (
                        id SERIAL PRIMARY KEY,
                        employee_id TEXT NOT NULL,
                        date DATE NOT NULL,
                        first_entry TIME,
                        last_exit TIME,
                        total_hours DECIMAL(4,2) DEFAULT 0,
                        worked_day BOOLEAN DEFAULT false,
                        is_holiday BOOLEAN DEFAULT false,
                        is_weekend BOOLEAN DEFAULT false,
                        late_minutes INTEGER DEFAULT 0,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        UNIQUE(employee_id, date)
                    )
                ''')
                
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS weekly_shift_assignments (
                        id SERIAL PRIMARY KEY,
                        employee_id TEXT NOT NULL,
                        week_start DATE NOT NULL,
                        week_end DATE NOT NULL,
                        shift_type TEXT NOT NULL,
                        start_time TIME NOT NULL,
                        end_time TIME NOT NULL,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        UNIQUE(employee_id, week_start)
                    )
                ''')
                
                # Crear índices para optimización
                cursor.execute('CREATE INDEX IF NOT EXISTS idx_attendance_employee_id ON attendance_records(employee_id)')
                cursor.execute('CREATE INDEX IF NOT EXISTS idx_attendance_timestamp ON attendance_records(timestamp)')
                cursor.execute('CREATE INDEX IF NOT EXISTS idx_attendance_date ON attendance_records(DATE(timestamp))')
                cursor.execute('CREATE INDEX IF NOT EXISTS idx_employees_active ON employees(active)')
                cursor.execute('CREATE INDEX IF NOT EXISTS idx_weekly_shifts_employee_week ON weekly_shift_assignments(employee_id, week_start)')
                
            else:
                # Crear tablas SQLite
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS employees (
                        id INTEGER PRIMARY KEY,
                        employee_id TEXT UNIQUE,
                        name TEXT,
                        department TEXT DEFAULT 'General',
                        schedule TEXT DEFAULT 'estandar',
                        phone TEXT DEFAULT '',
                        email TEXT DEFAULT '',
                        active BOOLEAN DEFAULT 1,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        synced_to_device BOOLEAN DEFAULT 0
                    )
                ''')
                
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS employee_schedules (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        employee_id TEXT NOT NULL,
                        schedule_type TEXT NOT NULL,
                        shift_type TEXT DEFAULT NULL,
                        start_time TEXT NOT NULL,
                        end_time TEXT NOT NULL,
                        days_of_week TEXT NOT NULL,
                        active_from DATE DEFAULT CURRENT_DATE,
                        active_until DATE DEFAULT NULL,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                    )
                ''')
                
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS attendance_records (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        employee_id TEXT,
                        event_type TEXT,
                        timestamp TIMESTAMP,
                        reader_no INTEGER DEFAULT 1,
                        verify_method TEXT DEFAULT 'huella',
                        status TEXT DEFAULT 'autorizado'
                    )
                ''')
                
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS daily_summaries (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        employee_id TEXT NOT NULL,
                        date DATE NOT NULL,
                        first_entry TEXT,
                        last_exit TEXT,
                        total_hours REAL DEFAULT 0,
                        worked_day BOOLEAN DEFAULT 0,
                        is_holiday BOOLEAN DEFAULT 0,
                        is_weekend BOOLEAN DEFAULT 0,
                        late_minutes INTEGER DEFAULT 0,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        UNIQUE(employee_id, date)
                    )
                ''')
                
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS weekly_shift_assignments (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        employee_id TEXT NOT NULL,
                        week_start DATE NOT NULL,
                        week_end DATE NOT NULL,
                        shift_type TEXT NOT NULL,
                        start_time TEXT NOT NULL,
                        end_time TEXT NOT NULL,
                        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                        UNIQUE(employee_id, week_start)
                    )
                ''')
                
                cursor.execute('''
                    INSERT OR IGNORE INTO employees (employee_id, name, department) 
                    VALUES (?, ?, ?)
                ''', ('1', 'Administrador', 'Administración'))
                cursor.execute('CREATE INDEX IF NOT EXISTS idx_attendance_employee_id ON attendance_records(employee_id)')
                cursor.execute('CREATE INDEX IF NOT EXISTS idx_attendance_timestamp ON attendance_records(timestamp)')
                cursor.execute('CREATE INDEX IF NOT EXISTS idx_attendance_date ON attendance_records(date(timestamp))')
                cursor.execute('CREATE INDEX IF NOT EXISTS idx_employees_active ON employees(active)')
                cursor.execute('CREATE INDEX IF NOT EXISTS idx_weekly_shifts_employee_week ON weekly_shift_assignments(employee_id, week_start)')
            
            conn.commit()
            print("Base de datos inicializada")
            
        except Exception as e:
            print(f"Error inicializando base de datos: {e}")
        finally:
            conn.close()
    
    def test_connection(self):
        """Probar conexión al dispositivo"""
        try:
            endpoints = [
                f"{self.base_url}/System/deviceInfo",
                f"http://{self.device_ip}/ISAPI/System/status",
                f"http://{self.device_ip}"
            ]
            
            for endpoint in endpoints:
                try:
                    response = self.session.get(endpoint, timeout=2)
                    if response.status_code in [200, 401]:
                        self.connected = True
                        return True
                except:
                    continue
                    
            self.connected = False
            return False
        except:
            self.connected = False
            return False
    
    def record_attendance(self, employee_id, timestamp, reader_no=1, verify_method="huella"):
        """Registrar asistencia"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # Obtener información del empleado
            if self.db_type == 'postgresql':
                cursor.execute('SELECT name, department, schedule FROM employees WHERE employee_id = %s', (employee_id,))
            else:
                cursor.execute('SELECT name, department, schedule FROM employees WHERE employee_id = ?', (employee_id,))
            
            employee = cursor.fetchone()
            
            if not employee:
                print(f"Empleado {employee_id} no encontrado")
                conn.close()
                return False
            
            # Usar hora local de Colombia directamente
            local_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Verificar duplicados
            if self.db_type == 'postgresql':
                cursor.execute('''
                    SELECT timestamp FROM attendance_records 
                    WHERE employee_id = %s AND timestamp > NOW() - INTERVAL '10 seconds'
                    ORDER BY timestamp DESC LIMIT 1
                ''', (employee_id,))
            else:
                cursor.execute('''
                    SELECT timestamp FROM attendance_records 
                    WHERE employee_id = ? AND datetime(timestamp) > datetime('now', '-10 seconds')
                    ORDER BY timestamp DESC LIMIT 1
                ''', (employee_id,))
            
            recent_record = cursor.fetchone()
            if recent_record:
                print(f"DUPLICADO EVITADO: {employee[0]} - Registro muy reciente")
                conn.close()
                return False
            
            # Determinar tipo de evento
            event_type = self.determine_event_type(employee_id)
            
            # Determinar si es break o almuerzo
            is_break = event_type.startswith('break_')
            is_lunch = event_type.startswith('almuerzo_')
            
            break_type = None
            if is_break:
                if employee[1] == 'Operativos':
                    break_type = 'operativo_break'
                else:
                    break_type = 'admin_break'
            elif is_lunch:
                break_type = 'almuerzo_admin'
            
            # Insertar registro
            if self.db_type == 'postgresql':
                cursor.execute('''
                    INSERT INTO attendance_records 
                    (employee_id, event_type, timestamp, reader_no, verify_method, status, is_break_record, break_type)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                ''', (employee_id, event_type, local_timestamp, reader_no, verify_method, "autorizado", is_break or is_lunch, break_type))
            else:
                cursor.execute('''
                    INSERT INTO attendance_records 
                    (employee_id, event_type, timestamp, reader_no, verify_method, status, is_break_record, break_type)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (employee_id, event_type, local_timestamp, reader_no, verify_method, "autorizado", is_break or is_lunch, break_type))
            
            conn.commit()
            conn.close()
            
            print(f"REGISTRO: {employee[0]} - {event_type.upper()} - {local_timestamp}")
            
            # Mostrar tipo de break o almuerzo si aplica
            if is_break:
                if employee[1] == 'Operativos':
                    break_display = 'BREAK OPERATIVO' if event_type == 'break_salida' else 'REGRESO DE BREAK OPERATIVO'
                else:
                    break_display = 'BREAK ADMINISTRATIVO' if event_type == 'break_salida' else 'REGRESO DE BREAK'
                print(f"BREAK: {break_display}")
            elif is_lunch:
                lunch_display = 'SALIDA A ALMUERZO' if event_type == 'almuerzo_salida' else 'REGRESO DE ALMUERZO'
                print(f"ALMUERZO: {lunch_display}")
            
            # Emitir evento WebSocket
            socketio.emit('attendance_record', {
                'employee_id': employee_id,
                'name': employee[0],
                'event_type': event_type,
                'timestamp': local_timestamp,
                'verify_method': verify_method,
                'department': employee[1] or 'General',
                'schedule': employee[2] or 'estandar',
                'real_time': True,
                'is_break': is_break,
                'is_lunch': is_lunch,
                'break_type': break_type
            })
            
            # Verificar tardanza solo para la primera entrada del día
            if event_type == 'entrada':
                self.check_late_arrival_first_entry(employee_id, employee[0], employee[1], employee[2], local_timestamp)
            
            # Actualizar resumen diario
            self.update_daily_summary(employee_id, local_timestamp.split(' ')[0])
            
            return True
            
        except Exception as e:
            print(f"Error al registrar: {e}")
            return False
    
    def determine_event_type(self, employee_id):
        """Determinar entrada, salida, break o almuerzo"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        # Obtener información del empleado y último registro
        if self.db_type == 'postgresql':
            cursor.execute('''
                SELECT e.department, ar.event_type FROM employees e
                LEFT JOIN (
                    SELECT employee_id, event_type FROM attendance_records 
                    WHERE employee_id = %s AND DATE(timestamp) = CURRENT_DATE
                    ORDER BY timestamp DESC LIMIT 1
                ) ar ON e.employee_id = ar.employee_id
                WHERE e.employee_id = %s
            ''', (employee_id, employee_id))
        else:
            cursor.execute('''
                SELECT e.department, ar.event_type FROM employees e
                LEFT JOIN (
                    SELECT employee_id, event_type FROM attendance_records 
                    WHERE employee_id = ? AND date(timestamp) = date('now')
                    ORDER BY timestamp DESC LIMIT 1
                ) ar ON e.employee_id = ar.employee_id
                WHERE e.employee_id = ?
            ''', (employee_id, employee_id))
        
        result = cursor.fetchone()
        
        # Obtener turno del empleado si es operativo
        shift_type = None
        if result and result[0] == 'Operativos':
            today = datetime.now().strftime('%Y-%m-%d')
            week_start = datetime.now().date() - timedelta(days=datetime.now().weekday())
            
            if self.db_type == 'postgresql':
                cursor.execute('''
                    SELECT shift_type FROM weekly_shift_assignments 
                    WHERE employee_id = %s AND week_start = %s
                ''', (employee_id, week_start))
            else:
                cursor.execute('''
                    SELECT shift_type FROM weekly_shift_assignments 
                    WHERE employee_id = ? AND week_start = ?
                ''', (employee_id, week_start))
            
            shift_result = cursor.fetchone()
            if shift_result:
                shift_type = shift_result[0]
        
        conn.close()
        
        if not result:
            return 'entrada'
        
        department, last_event = result
        current_time = datetime.now().time()
        
        # FASE 2: Detección de almuerzo para departamentos administrativos (12:00-14:00)
        if department in ['Reacondicionamiento', 'Logistica', 'Administracion']:
            # Almuerzo: 12:00-14:00
            if dt_time(12, 0) <= current_time <= dt_time(14, 0):
                if last_event in ['entrada', 'break_entrada']:
                    return 'almuerzo_salida'
                elif last_event == 'almuerzo_salida':
                    return 'almuerzo_entrada'
            
            # Break de mañana: 9:00-10:00
            elif dt_time(9, 0) <= current_time <= dt_time(10, 0):
                if last_event == 'entrada':
                    return 'break_salida'
                elif last_event == 'break_salida':
                    return 'break_entrada'
        
        # FASE 3: Breaks para operativos por turnos
        elif department == 'Operativos' and shift_type:
            if shift_type == 'mañana' and dt_time(9, 0) <= current_time <= dt_time(10, 0):
                if last_event == 'entrada':
                    return 'break_salida'
                elif last_event == 'break_salida':
                    return 'break_entrada'
            elif shift_type == 'tarde' and dt_time(17, 0) <= current_time <= dt_time(18, 0):
                if last_event == 'entrada':
                    return 'break_salida'
                elif last_event == 'break_salida':
                    return 'break_entrada'
            elif shift_type == 'noche' and dt_time(1, 0) <= current_time <= dt_time(2, 0):
                if last_event == 'entrada':
                    return 'break_salida'
                elif last_event == 'break_salida':
                    return 'break_entrada'
        
        # Lógica normal entrada/salida
        if not last_event:
            return 'entrada'
        
        return 'salida' if last_event in ['entrada', 'break_entrada', 'almuerzo_entrada'] else 'entrada'
    
    def add_employee(self, employee_id, name, department="General", schedule="estandar", phone="", email=""):
        """Agregar empleado"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            if self.db_type == 'postgresql':
                cursor.execute('''
                    INSERT INTO employees (employee_id, name, department, schedule, phone, email) 
                    VALUES (%s, %s, %s, %s, %s, %s)
                ''', (employee_id, name, department, schedule, phone, email))
            else:
                cursor.execute('''
                    INSERT INTO employees (employee_id, name, department, schedule, phone, email) 
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (employee_id, name, department, schedule, phone, email))
            
            conn.commit()
            conn.close()
            
            # Limpiar cache
            self.employees_cache = {}
            self.cache_timestamp = 0
            
            socketio.emit('employee_added', {
                'employee_id': employee_id,
                'name': name,
                'department': department,
                'schedule': schedule
            })
            
            return True, f"Empleado {name} agregado exitosamente"
            
        except Exception as e:
            conn.close()
            if "unique" in str(e).lower() or "duplicate" in str(e).lower():
                return False, f"El empleado con ID {employee_id} ya existe"
            return False, f"Error: {str(e)}"
    
    def get_employees(self):
        """Obtener lista de empleados con cache"""
        current_time = time_module.time()
        
        # Usar cache si está vigente
        if (current_time - self.cache_timestamp) < self.cache_duration and self.employees_cache:
            return list(self.employees_cache.values())
        
        conn = self.get_connection()
        cursor = conn.cursor()
        
        cursor.execute('''
            SELECT employee_id, name, department, schedule, phone, email, active, synced_to_device, created_at
            FROM employees ORDER BY name
        ''')
        
        employees = cursor.fetchall()
        conn.close()
        
        # Actualizar cache
        self.employees_cache = {}
        for emp in employees:
            employee_data = {
                'employee_id': emp[0], 'name': emp[1], 'department': emp[2] or 'General', 
                'schedule': emp[3] or 'estandar', 'phone': emp[4] or '', 'email': emp[5] or '',
                'active': emp[6], 'synced': emp[7], 'created_at': emp[8]
            }
            self.employees_cache[emp[0]] = employee_data
        
        self.cache_timestamp = current_time
        return list(self.employees_cache.values())
    
    def calculate_worked_hours(self, entrada_time, salida_time, department):
        """Calcular horas reales trabajadas descontando breaks/almuerzos"""
        if not entrada_time or not salida_time:
            return 0
        
        try:
            # Convertir a datetime para cálculos
            if isinstance(entrada_time, str):
                entrada_dt = datetime.strptime(entrada_time, '%H:%M:%S')
            else:
                entrada_dt = datetime.combine(datetime.today(), entrada_time)
            
            if isinstance(salida_time, str):
                salida_dt = datetime.strptime(salida_time, '%H:%M:%S')
            else:
                salida_dt = datetime.combine(datetime.today(), salida_time)
            
            # Si la salida es al día siguiente (turno nocturno)
            if salida_dt.time() < entrada_dt.time():
                salida_dt = salida_dt + timedelta(days=1)
            
            # Calcular horas brutas
            total_seconds = (salida_dt - entrada_dt).total_seconds()
            hours_worked = total_seconds / 3600
            
            # Descontar breaks según departamento
            if department in ['Reacondicionamiento', 'Logistica', 'Administracion']:
                # Descontar 20 min de break + 60 min de almuerzo = 80 min = 1.33 horas
                hours_worked -= 80 / 60
            elif department == 'Operativos':
                # Descontar solo 20 min de break = 0.33 horas
                hours_worked -= 20 / 60
            else:
                # Otros departamentos: descontar break + almuerzo
                hours_worked -= 80 / 60
            
            return max(0, round(hours_worked, 2))
            
        except Exception as e:
            print(f"Error calculando horas: {e}")
            return 0
    
    def get_dashboard_data(self):
        """Obtener datos del dashboard"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            # Registros de hoy
            if self.db_type == 'postgresql':
                cursor.execute("SELECT COUNT(*) FROM attendance_records WHERE DATE(timestamp) = CURRENT_DATE")
                total_records = cursor.fetchone()[0]
                
                cursor.execute("SELECT COUNT(DISTINCT employee_id) FROM attendance_records WHERE DATE(timestamp) = CURRENT_DATE")
                unique_employees = cursor.fetchone()[0]
                
                # Estado de empleados
                cursor.execute('''
                    SELECT e.name, e.employee_id, ar.event_type, ar.timestamp
                    FROM employees e
                    LEFT JOIN (
                        SELECT DISTINCT ON (employee_id) employee_id, event_type, timestamp
                        FROM attendance_records
                        WHERE DATE(timestamp) = CURRENT_DATE
                        ORDER BY employee_id, timestamp DESC
                    ) ar ON e.employee_id = ar.employee_id
                    WHERE e.active = true
                ''')
                employees_status = cursor.fetchall()
                
                # Registros recientes (solo empleados activos)
                cursor.execute('''
                    SELECT e.name, ar.event_type, ar.timestamp, ar.verify_method
                    FROM attendance_records ar
                    JOIN employees e ON ar.employee_id = e.employee_id
                    WHERE e.active = true
                    ORDER BY ar.timestamp DESC LIMIT 20
                ''')
                recent_records = cursor.fetchall()
                
            else:
                cursor.execute("SELECT COUNT(*) FROM attendance_records WHERE date(timestamp) = date('now')")
                total_records = cursor.fetchone()[0]
                
                cursor.execute("SELECT COUNT(DISTINCT employee_id) FROM attendance_records WHERE date(timestamp) = date('now')")
                unique_employees = cursor.fetchone()[0]
                
                cursor.execute('''
                    SELECT e.name, e.employee_id, ar.event_type, ar.timestamp
                    FROM employees e
                    LEFT JOIN (
                        SELECT employee_id, event_type, timestamp,
                               ROW_NUMBER() OVER (PARTITION BY employee_id ORDER BY timestamp DESC) as rn
                        FROM attendance_records
                        WHERE date(timestamp) = date('now')
                    ) ar ON e.employee_id = ar.employee_id AND ar.rn = 1
                    WHERE e.active = 1
                ''')
                employees_status = cursor.fetchall()
                
                cursor.execute('''
                    SELECT e.name, ar.event_type, ar.timestamp, ar.verify_method
                    FROM attendance_records ar
                    JOIN employees e ON ar.employee_id = e.employee_id
                    WHERE e.active = 1
                    ORDER BY ar.timestamp DESC LIMIT 20
                ''')
                recent_records = cursor.fetchall()
            
            conn.close()
            
            # Procesar estado de empleados
            inside = []
            outside = []
            
            for emp in employees_status:
                name, emp_id, last_event, timestamp = emp
                if last_event == 'entrada':
                    inside.append({'name': name, 'id': emp_id, 'time': timestamp})
                else:
                    outside.append({'name': name, 'id': emp_id, 'time': timestamp})
            
            return {
                'total_records': total_records,
                'unique_employees': unique_employees,
                'employees_inside': inside,
                'employees_outside': outside,
                'recent_records': recent_records,
                'connected': self.connected,
                'monitoring': self.monitoring
            }
            
        except Exception as e:
            print(f"Error obteniendo datos dashboard: {e}")
            conn.close()
            return {
                'total_records': 0,
                'unique_employees': 0,
                'employees_inside': [],
                'employees_outside': [],
                'recent_records': [],
                'connected': self.connected,
                'monitoring': self.monitoring
            }
    
    def is_work_day(self, date_obj, schedule, department):
        """Determinar si es día laboral según horarios por departamento"""
        day_of_week = date_obj.weekday()  # 0=Lunes, 6=Domingo
        
        # Horarios por departamento
        if department in ['Reacondicionamiento', 'Logistica', 'Administracion']:
            # L-J: 7:00-17:00, V: 7:00-16:00, S-D: No laboral
            return day_of_week < 5  # Lunes a Viernes
        elif department == 'Operativos':
            # Todos los días según turno asignado
            return True
        else:
            # General: L-V: 8:00-17:00
            return day_of_week < 5  # Lunes a Viernes
    
    def get_expected_hours_by_department(self, department, day_of_week):
        """Obtener horarios esperados por departamento y día"""
        if department in ['Reacondicionamiento', 'Logistica', 'Administracion']:
            if day_of_week < 4:  # Lunes a Jueves
                return ('07:00', '17:00')
            elif day_of_week == 4:  # Viernes
                return ('07:00', '16:00')
            else:  # Fin de semana
                return None
        elif department == 'Operativos':
            # Depende del turno asignado - se manejará por separado
            return ('06:00', '14:00')  # Default mañana
        else:
            # General
            if day_of_week < 5:  # Lunes a Viernes
                return ('08:00', '17:00')
            else:
                return None
    
    def generate_attendance_report(self, start_date, end_date, employee_id=None, department=None):
        """Generar reporte de asistencia mejorado con cálculos precisos"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            # Obtener todos los empleados activos
            if self.db_type == 'postgresql':
                emp_query = 'SELECT employee_id, name, department, schedule FROM employees WHERE active = true'
            else:
                emp_query = 'SELECT employee_id, name, department, schedule FROM employees WHERE active = 1'
            
            emp_params = []
            if employee_id:
                emp_query += ' AND employee_id = {}'
                emp_query = emp_query.format('%s' if self.db_type == 'postgresql' else '?')
                emp_params.append(employee_id)
            
            if department and department != 'General':
                emp_query += ' AND department = {}'
                emp_query = emp_query.format('%s' if self.db_type == 'postgresql' else '?')
                emp_params.append(department)
            
            emp_query += ' ORDER BY name'
            cursor.execute(emp_query, emp_params)
            employees = cursor.fetchall()
            
            if not employees:
                conn.close()
                return {}
            
            # Obtener registros de asistencia
            if self.db_type == 'postgresql':
                records_query = '''
                    SELECT employee_id, event_type, timestamp FROM attendance_records 
                    WHERE DATE(timestamp) BETWEEN %s AND %s AND event_type IN ('entrada', 'salida')
                    ORDER BY employee_id, timestamp
                '''
            else:
                records_query = '''
                    SELECT employee_id, event_type, timestamp FROM attendance_records 
                    WHERE date(timestamp) BETWEEN ? AND ? AND event_type IN ('entrada', 'salida')
                    ORDER BY employee_id, timestamp
                '''
            
            cursor.execute(records_query, [start_date, end_date])
            records = cursor.fetchall()
            conn.close()
            
            # Procesar datos por empleado
            report_data = {}
            current_date = datetime.strptime(start_date, '%Y-%m-%d')
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
            
            # Inicializar estructura para cada empleado
            for emp in employees:
                emp_id, name, dept, schedule = emp
                report_data[emp_id] = {
                    'name': name,
                    'department': dept or 'General',
                    'schedule': schedule or 'general',
                    'summary': {
                        'total_days_worked': 0,
                        'total_hours': 0,
                        'late_days': 0,
                        'absent_days': 0,
                        'weekend_days': 0,
                        'average_daily_hours': 0
                    },
                    'days': {}
                }
            
            # Llenar todos los días del rango
            temp_date = current_date
            while temp_date <= end_date_obj:
                date_str = temp_date.strftime('%Y-%m-%d')
                day_name = temp_date.strftime('%A')
                day_of_week = temp_date.weekday()
                
                for emp_id in report_data:
                    emp_data = report_data[emp_id]
                    department = emp_data['department']
                    
                    # Determinar si es día laboral
                    is_work_day = self.is_work_day(temp_date, emp_data['schedule'], department)
                    expected_hours = self.get_expected_hours_by_department(department, day_of_week) if is_work_day else None
                    
                    emp_data['days'][date_str] = {
                        'date': temp_date.strftime('%d/%m/%Y'),
                        'day_name': day_name,
                        'expected_hours': expected_hours,
                        'entrada': None,
                        'salida': None,
                        'status': 'No laborable' if not is_work_day else 'Ausente',
                        'late': False,
                        'early_exit': False,
                        'hours_worked': 0,
                        'late_minutes': 0,
                        'early_minutes': 0,
                        'observations': []
                    }
                    
                    if not is_work_day:
                        emp_data['summary']['weekend_days'] += 1
                
                temp_date += timedelta(days=1)
            
            # Procesar registros de asistencia
            for record in records:
                emp_id, event_type, timestamp = record
                if emp_id not in report_data:
                    continue
                
                # Extraer fecha y hora
                timestamp_str = str(timestamp)
                if 'T' in timestamp_str:
                    date_part = timestamp_str.split('T')[0]
                    time_part = timestamp_str.split('T')[1][:8]
                else:
                    date_part = timestamp_str[:10]
                    time_part = timestamp_str[11:19]
                
                if date_part in report_data[emp_id]['days']:
                    day_data = report_data[emp_id]['days'][date_part]
                    
                    if event_type == 'entrada':
                        if not day_data['entrada']:
                            day_data['entrada'] = time_part
                    elif event_type == 'salida':
                        day_data['salida'] = time_part  # Última salida
            
            # Calcular estadísticas finales
            for emp_id in report_data:
                emp_data = report_data[emp_id]
                department = emp_data['department']
                
                for date_str, day_data in emp_data['days'].items():
                    if day_data['expected_hours'] is None:  # No laborable
                        continue
                    
                    entrada = day_data['entrada']
                    salida = day_data['salida']
                    expected_start, expected_end = day_data['expected_hours']
                    
                    if entrada and salida:
                        # Calcular horas trabajadas con descuentos
                        hours_worked = self.calculate_worked_hours(entrada, salida, department)
                        day_data['hours_worked'] = hours_worked
                        day_data['status'] = 'Presente'
                        
                        # Verificar tardanza
                        entrada_time = datetime.strptime(entrada, '%H:%M:%S').time()
                        expected_start_time = datetime.strptime(expected_start, '%H:%M').time()
                        
                        if entrada_time > expected_start_time:
                            day_data['late'] = True
                            entrada_dt = datetime.combine(datetime.today(), entrada_time)
                            expected_dt = datetime.combine(datetime.today(), expected_start_time)
                            day_data['late_minutes'] = int((entrada_dt - expected_dt).total_seconds() / 60)
                            day_data['observations'].append(f"Tardó {day_data['late_minutes']} min")
                            emp_data['summary']['late_days'] += 1
                        
                        # Verificar salida temprana
                        salida_time = datetime.strptime(salida, '%H:%M:%S').time()
                        expected_end_time = datetime.strptime(expected_end, '%H:%M').time()
                        
                        if salida_time < expected_end_time:
                            day_data['early_exit'] = True
                            salida_dt = datetime.combine(datetime.today(), salida_time)
                            expected_dt = datetime.combine(datetime.today(), expected_end_time)
                            day_data['early_minutes'] = int((expected_dt - salida_dt).total_seconds() / 60)
                            day_data['observations'].append(f"Salió {day_data['early_minutes']} min temprano")
                        
                        emp_data['summary']['total_days_worked'] += 1
                        emp_data['summary']['total_hours'] += hours_worked
                        
                    elif entrada:
                        day_data['status'] = 'Sin salida'
                        day_data['observations'].append('Falta registro de salida')
                    elif salida:
                        day_data['status'] = 'Sin entrada'
                        day_data['observations'].append('Falta registro de entrada')
                    else:
                        emp_data['summary']['absent_days'] += 1
                
                # Calcular promedio de horas diarias
                if emp_data['summary']['total_days_worked'] > 0:
                    emp_data['summary']['average_daily_hours'] = round(
                        emp_data['summary']['total_hours'] / emp_data['summary']['total_days_worked'], 2
                    )
            
            return report_data
            
        except Exception as e:
            conn.close()
            print(f"Error generando reporte: {e}")
            return {}
    
    def export_to_excel(self, report_data, filename):
        """Exportar reporte a Excel"""
        if not EXCEL_AVAILABLE:
            raise Exception("openpyxl no está instalado")
            
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Asistencia"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        present_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        absent_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        late_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        # Encabezados
        headers = ['Empleado', 'Departamento', 'Fecha', 'Día', 'Horario', 'Entrada', 'Salida', 'Horas', 'Estado', 'Observaciones']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        row = 2
        for emp_id, emp_data in report_data.items():
            for date_str, day_data in sorted(emp_data['days'].items()):
                ws.cell(row=row, column=1, value=emp_data['name'])
                ws.cell(row=row, column=2, value=emp_data['department'])
                ws.cell(row=row, column=3, value=day_data['date'])
                ws.cell(row=row, column=4, value=day_data['day_name'])
                
                if day_data['expected_hours']:
                    ws.cell(row=row, column=5, value=f"{day_data['expected_hours'][0]} - {day_data['expected_hours'][1]}")
                else:
                    ws.cell(row=row, column=5, value="No laborable")
                
                ws.cell(row=row, column=6, value=str(day_data['entrada']) if day_data['entrada'] else "")
                ws.cell(row=row, column=7, value=str(day_data['salida']) if day_data['salida'] else "")
                ws.cell(row=row, column=8, value=day_data['hours_worked'])
                ws.cell(row=row, column=9, value=day_data['status'])
                
                # Observaciones
                obs = []
                if day_data['late']:
                    obs.append("Tardó")
                if day_data['early_exit']:
                    obs.append("Salió temprano")
                ws.cell(row=row, column=10, value=", ".join(obs))
                
                # Colores según estado
                if day_data['status'] == 'Presente':
                    if day_data['late']:
                        for col in range(1, 11):
                            ws.cell(row=row, column=col).fill = late_fill
                    else:
                        for col in range(1, 11):
                            ws.cell(row=row, column=col).fill = present_fill
                elif day_data['status'] == 'Ausente':
                    for col in range(1, 11):
                        ws.cell(row=row, column=col).fill = absent_fill
                
                row += 1
        
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Guardar archivo
        wb.save(filename)
        return filename
    
    def export_to_pdf(self, report_data, filename):
        """Exportar reporte a PDF"""
        if not PDF_AVAILABLE:
            raise Exception("reportlab no está instalado")
            
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Título
        title = Paragraph("<b>REPORTE DE ASISTENCIA - PCSHEK</b>", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Crear tabla de datos
        data = [['Empleado', 'Depto', 'Fecha', 'Entrada', 'Salida', 'Horas', 'Estado']]
        
        for emp_id, emp_data in report_data.items():
            for date_str, day_data in sorted(emp_data['days'].items()):
                row = [
                    emp_data['name'][:15],
                    emp_data['department'][:8],
                    day_data['date'],
                    str(day_data['entrada'])[:5] if day_data['entrada'] else "-",
                    str(day_data['salida'])[:5] if day_data['salida'] else "-",
                    str(day_data['hours_worked']),
                    day_data['status'][:10]
                ]
                data.append(row)
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        story.append(table)
        doc.build(story)
        
        with open(filename, 'wb') as f:
            f.write(buffer.getvalue())
        
        return filename
    
    def check_late_arrival_first_entry(self, employee_id, name, department, schedule, timestamp):
        """Verificar si la primera entrada del día es tardía usando horarios por departamento"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # Verificar si es la primera entrada del día
            today = timestamp.split(' ')[0]
            
            if self.db_type == 'postgresql':
                cursor.execute('''
                    SELECT COUNT(*) FROM attendance_records 
                    WHERE employee_id = %s AND DATE(timestamp) = %s AND event_type = 'entrada'
                ''', (employee_id, today))
            else:
                cursor.execute('''
                    SELECT COUNT(*) FROM attendance_records 
                    WHERE employee_id = ? AND date(timestamp) = ? AND event_type = 'entrada'
                ''', (employee_id, today))
            
            entry_count = cursor.fetchone()[0]
            conn.close()
            
            # Solo verificar tardanza si es la primera entrada del día
            if entry_count > 1:
                return
            
            # Obtener horario esperado por departamento
            arrival_time = datetime.strptime(timestamp, '%Y-%m-%d %H:%M:%S')
            day_of_week = arrival_time.weekday()
            
            expected_hours = self.get_expected_hours_by_department(department, day_of_week)
            if not expected_hours:
                return  # No es día laboral
            
            # Verificar tardanza
            arrival_time_only = arrival_time.time()
            expected_start = datetime.strptime(expected_hours[0], '%H:%M').time()
            
            if arrival_time_only > expected_start:
                # Calcular minutos de tardanza
                arrival_dt = datetime.combine(datetime.today(), arrival_time_only)
                expected_dt = datetime.combine(datetime.today(), expected_start)
                late_minutes = int((arrival_dt - expected_dt).total_seconds() / 60)
                
                # Emitir notificación de tardanza
                socketio.emit('late_arrival_alert', {
                    'employee_id': employee_id,
                    'name': name,
                    'department': department,
                    'expected_time': expected_hours[0],
                    'actual_time': arrival_time_only.strftime('%H:%M'),
                    'late_minutes': late_minutes,
                    'timestamp': timestamp,
                    'severity': 'severe' if late_minutes > 30 else 'moderate' if late_minutes > 15 else 'mild'
                })
                
                print(f"TARDANZA: {name} llegó {late_minutes} minutos tarde")
                
        except Exception as e:
            print(f"Error verificando tardanza: {e}")
    
    def update_daily_summary(self, employee_id, date):
        """Actualizar resumen diario del empleado con cálculos mejorados"""
        try:
            conn = self.get_connection()
            cursor = conn.cursor()
            
            # Obtener información del empleado
            if self.db_type == 'postgresql':
                cursor.execute('SELECT name, department FROM employees WHERE employee_id = %s', (employee_id,))
            else:
                cursor.execute('SELECT name, department FROM employees WHERE employee_id = ?', (employee_id,))
            
            employee_info = cursor.fetchone()
            if not employee_info:
                conn.close()
                return
            
            name, department = employee_info
            
            # Obtener todos los registros del día (solo entrada/salida)
            if self.db_type == 'postgresql':
                cursor.execute('''
                    SELECT event_type, timestamp FROM attendance_records 
                    WHERE employee_id = %s AND DATE(timestamp) = %s AND event_type IN ('entrada', 'salida')
                    ORDER BY timestamp
                ''', (employee_id, date))
            else:
                cursor.execute('''
                    SELECT event_type, timestamp FROM attendance_records 
                    WHERE employee_id = ? AND date(timestamp) = ? AND event_type IN ('entrada', 'salida')
                    ORDER BY timestamp
                ''', (employee_id, date))
            
            records = cursor.fetchall()
            
            if not records:
                conn.close()
                return
            
            # Calcular primera entrada y última salida
            first_entry = None
            last_exit = None
            
            for event_type, timestamp in records:
                time_part = str(timestamp)[11:19] if 'T' in str(timestamp) else str(timestamp).split(' ')[1][:8]
                
                if event_type == 'entrada':
                    if not first_entry:
                        first_entry = time_part
                elif event_type == 'salida':
                    last_exit = time_part
            
            # Calcular horas trabajadas con descuentos
            total_hours = 0
            worked_day = False
            
            if first_entry and last_exit:
                total_hours = self.calculate_worked_hours(first_entry, last_exit, department)
                worked_day = total_hours > 1  # Mínimo 1 hora para contar como día trabajado
            
            # Verificar si es fin de semana
            date_obj = datetime.strptime(date, '%Y-%m-%d')
            is_weekend = not self.is_work_day(date_obj, 'general', department)
            
            # Insertar o actualizar resumen
            if self.db_type == 'postgresql':
                cursor.execute('''
                    INSERT INTO daily_summaries 
                    (employee_id, date, first_entry, last_exit, total_hours, worked_day, is_weekend)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (employee_id, date) 
                    DO UPDATE SET 
                        first_entry = EXCLUDED.first_entry,
                        last_exit = EXCLUDED.last_exit,
                        total_hours = EXCLUDED.total_hours,
                        worked_day = EXCLUDED.worked_day
                ''', (employee_id, date, first_entry, last_exit, total_hours, worked_day, is_weekend))
            else:
                cursor.execute('''
                    INSERT OR REPLACE INTO daily_summaries 
                    (employee_id, date, first_entry, last_exit, total_hours, worked_day, is_weekend)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (employee_id, date, first_entry, last_exit, total_hours, worked_day, is_weekend))
            
            conn.commit()
            conn.close()
            
        except Exception as e:
            print(f"Error actualizando resumen diario: {e}")
        """Obtener datos del dashboard"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        try:
            # Registros de hoy
            if self.db_type == 'postgresql':
                cursor.execute("SELECT COUNT(*) FROM attendance_records WHERE DATE(timestamp) = CURRENT_DATE")
                total_records = cursor.fetchone()[0]
                
                cursor.execute("SELECT COUNT(DISTINCT employee_id) FROM attendance_records WHERE DATE(timestamp) = CURRENT_DATE")
                unique_employees = cursor.fetchone()[0]
                
                # Estado de empleados
                cursor.execute('''
                    SELECT e.name, e.employee_id, ar.event_type, ar.timestamp
                    FROM employees e
                    LEFT JOIN (
                        SELECT DISTINCT ON (employee_id) employee_id, event_type, timestamp
                        FROM attendance_records
                        WHERE DATE(timestamp) = CURRENT_DATE
                        ORDER BY employee_id, timestamp DESC
                    ) ar ON e.employee_id = ar.employee_id
                    WHERE e.active = true
                ''')
                employees_status = cursor.fetchall()
                
                # Registros recientes (solo empleados activos)
                cursor.execute('''
                    SELECT e.name, ar.event_type, ar.timestamp, ar.verify_method
                    FROM attendance_records ar
                    JOIN employees e ON ar.employee_id = e.employee_id
                    WHERE e.active = true
                    ORDER BY ar.timestamp DESC LIMIT 20
                ''')
                recent_records = cursor.fetchall()
                
            else:
                cursor.execute("SELECT COUNT(*) FROM attendance_records WHERE date(timestamp) = date('now')")
                total_records = cursor.fetchone()[0]
                
                cursor.execute("SELECT COUNT(DISTINCT employee_id) FROM attendance_records WHERE date(timestamp) = date('now')")
                unique_employees = cursor.fetchone()[0]
                
                cursor.execute('''
                    SELECT e.name, e.employee_id, ar.event_type, ar.timestamp
                    FROM employees e
                    LEFT JOIN (
                        SELECT employee_id, event_type, timestamp,
                               ROW_NUMBER() OVER (PARTITION BY employee_id ORDER BY timestamp DESC) as rn
                        FROM attendance_records
                        WHERE date(timestamp) = date('now')
                    ) ar ON e.employee_id = ar.employee_id AND ar.rn = 1
                    WHERE e.active = 1
                ''')
                employees_status = cursor.fetchall()
                
                cursor.execute('''
                    SELECT e.name, ar.event_type, ar.timestamp, ar.verify_method
                    FROM attendance_records ar
                    JOIN employees e ON ar.employee_id = e.employee_id
                    WHERE e.active = 1
                    ORDER BY ar.timestamp DESC LIMIT 20
                ''')
                recent_records = cursor.fetchall()
            
            conn.close()
            
            # Procesar estado de empleados
            inside = []
            outside = []
            
            for emp in employees_status:
                name, emp_id, last_event, timestamp = emp
                if last_event == 'entrada':
                    inside.append({'name': name, 'id': emp_id, 'time': timestamp})
                else:
                    outside.append({'name': name, 'id': emp_id, 'time': timestamp})
            
            return {
                'total_records': total_records,
                'unique_employees': unique_employees,
                'employees_inside': inside,
                'employees_outside': outside,
                'recent_records': recent_records,
                'connected': self.connected,
                'monitoring': self.monitoring
            }
            
        except Exception as e:
            print(f"Error obteniendo datos dashboard: {e}")
            conn.close()
            return {
                'total_records': 0,
                'unique_employees': 0,
                'employees_inside': [],
                'employees_outside': [],
                'recent_records': [],
                'connected': self.connected,
                'monitoring': self.monitoring
            }
    
    def start_monitoring(self):
        """Iniciar monitoreo"""
        if not self.monitoring:
            self.monitoring = True
            monitor_thread = threading.Thread(target=self._monitor_events, daemon=True)
            monitor_thread.start()
            print("Monitoreo iniciado")
    
    def stop_monitoring(self):
        """Detener monitoreo"""
        self.monitoring = False
        print("Monitoreo detenido")
    
    def _monitor_events(self):
        """Monitoreo de eventos del dispositivo"""
        url = f"http://{self.device_ip}/ISAPI/Event/notification/alertStream"
        
        while self.monitoring:
            try:
                response = self.session.get(url, stream=True, timeout=60)
                
                if response.status_code == 200:
                    self.connected = True
                    print("Stream de eventos activo")
                    
                    buffer = ""
                    for chunk in response.iter_content(chunk_size=1024):
                        if not self.monitoring:
                            break
                            
                        if chunk:
                            try:
                                chunk_str = chunk.decode('utf-8', errors='ignore')
                                buffer += chunk_str
                            except:
                                continue
                            
                            # Procesar eventos JSON
                            while '{' in buffer and '}' in buffer:
                                start = buffer.find('{')
                                if start == -1:
                                    break
                                    
                                brace_count = 0
                                end = start
                                
                                for i in range(start, len(buffer)):
                                    if buffer[i] == '{':
                                        brace_count += 1
                                    elif buffer[i] == '}':
                                        brace_count -= 1
                                        if brace_count == 0:
                                            end = i
                                            break
                                
                                if brace_count == 0:
                                    json_str = buffer[start:end+1]
                                    buffer = buffer[end+1:]
                                    
                                    try:
                                        event = json.loads(json_str)
                                        self._process_event(event)
                                    except json.JSONDecodeError:
                                        pass
                                else:
                                    break
                else:
                    self.connected = False
                    print(f"Error HTTP {response.status_code}")
                    time_module.sleep(10)
                    
            except Exception as e:
                self.connected = False
                print(f"Error monitoreo: {str(e)[:50]}...")
                time_module.sleep(5)
    
    def _process_event(self, event):
        """Procesar eventos del dispositivo"""
        if 'AccessControllerEvent' in event:
            acs_event = event['AccessControllerEvent']
            sub_type = acs_event.get('subEventType')
            
            if sub_type == 38:  # Acceso autorizado
                employee_id = acs_event.get('employeeNoString')
                timestamp = event.get('dateTime', datetime.now().isoformat())
                reader_no = acs_event.get('cardReaderNo', 1)
                verify_method = 'huella'  # Simplificado
                
                if employee_id:
                    self.record_attendance(employee_id, timestamp, reader_no, verify_method)

# Instancia global
system = OptimizedAttendanceSystem()

# WebSocket events
@socketio.on('connect')
def handle_connect():
    print('Cliente conectado')
    emit('status', {'connected': system.connected, 'monitoring': system.monitoring})

@socketio.on('disconnect')
def handle_disconnect():
    print('Cliente desconectado')

# Rutas web
@app.route('/')
def dashboard():
    return render_template('dashboard_pcshek.html')

@app.route('/employees')
def employees_page():
    return render_template('employees_pcshek.html')

@app.route('/employees/simple')
def employees_simple():
    return render_template('employees_simple.html')

@app.route('/api/dashboard')
def api_dashboard():
    return jsonify(system.get_dashboard_data())

@app.route('/api/employees')
def api_employees():
    return jsonify(system.get_employees())

@app.route('/api/employees', methods=['POST'])
def api_add_employee():
    data = request.json
    
    # Validación de datos
    if not data.get('employee_id') or not data.get('name'):
        return jsonify({'success': False, 'message': 'ID y nombre son obligatorios'})
    
    # Validar formato de email si se proporciona
    email = data.get('email', '').strip()
    if email and '@' not in email:
        return jsonify({'success': False, 'message': 'Formato de email inválido'})
    
    success, message = system.add_employee(
        data['employee_id'].strip(),
        data['name'].strip(),
        data.get('department', 'General'),
        data.get('schedule', 'estandar'),
        data.get('phone', '').strip(),
        email
    )
    return jsonify({'success': success, 'message': message})

@app.route('/api/start_monitoring', methods=['POST'])
def api_start_monitoring():
    system.start_monitoring()
    return jsonify({'success': True, 'message': 'Monitoreo iniciado'})

@app.route('/api/stop_monitoring', methods=['POST'])
def api_stop_monitoring():
    system.stop_monitoring()
    return jsonify({'success': True, 'message': 'Monitoreo detenido'})

@app.route('/api/employees/<employee_id>/toggle', methods=['POST'])
def api_toggle_employee(employee_id):
    conn = system.get_connection()
    cursor = conn.cursor()
    
    try:
        if system.db_type == 'postgresql':
            cursor.execute('SELECT name, active FROM employees WHERE employee_id = %s', (employee_id,))
        else:
            cursor.execute('SELECT name, active FROM employees WHERE employee_id = ?', (employee_id,))
        
        employee = cursor.fetchone()
        
        if not employee:
            conn.close()
            return jsonify({'success': False, 'message': 'Empleado no encontrado'})
        
        new_status = not employee[1]
        
        if system.db_type == 'postgresql':
            cursor.execute('UPDATE employees SET active = %s WHERE employee_id = %s', (new_status, employee_id))
        else:
            cursor.execute('UPDATE employees SET active = ? WHERE employee_id = ?', (new_status, employee_id))
        
        conn.commit()
        conn.close()
        
        # Limpiar cache
        system.employees_cache = {}
        system.cache_timestamp = 0
        
        status_text = "activado" if new_status else "desactivado"
        return jsonify({'success': True, 'message': f"Empleado {employee[0]} {status_text}"})
        
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'message': f"Error: {str(e)}"})

@app.route('/api/employees/<employee_id>', methods=['PUT'])
def api_update_employee(employee_id):
    data = request.json
    conn = system.get_connection()
    cursor = conn.cursor()
    
    try:
        if system.db_type == 'postgresql':
            cursor.execute('''
                UPDATE employees 
                SET name=%s, department=%s, phone=%s, email=%s, active=%s
                WHERE employee_id=%s
            ''', (data.get('name'), data.get('department'), data.get('phone', ''), 
                  data.get('email', ''), data.get('active', True), employee_id))
        else:
            cursor.execute('''
                UPDATE employees 
                SET name=?, department=?, phone=?, email=?, active=?
                WHERE employee_id=?
            ''', (data.get('name'), data.get('department'), data.get('phone', ''), 
                  data.get('email', ''), data.get('active', True), employee_id))
        
        if cursor.rowcount > 0:
            conn.commit()
            conn.close()
            
            # Limpiar cache
            system.employees_cache = {}
            system.cache_timestamp = 0
            
            return jsonify({'success': True, 'message': f"Empleado {data.get('name')} actualizado exitosamente"})
        else:
            conn.close()
            return jsonify({'success': False, 'message': 'Empleado no encontrado'})
            
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'message': f"Error: {str(e)}"})

@app.route('/api/employees/<employee_id>', methods=['DELETE'])
def api_delete_employee(employee_id):
    conn = system.get_connection()
    cursor = conn.cursor()
    
    try:
        # Obtener nombre antes de eliminar
        if system.db_type == 'postgresql':
            cursor.execute('SELECT name FROM employees WHERE employee_id = %s', (employee_id,))
        else:
            cursor.execute('SELECT name FROM employees WHERE employee_id = ?', (employee_id,))
        
        employee = cursor.fetchone()
        
        if not employee:
            conn.close()
            return jsonify({'success': False, 'message': 'Empleado no encontrado'})
        
        # Eliminar empleado
        if system.db_type == 'postgresql':
            cursor.execute('DELETE FROM employees WHERE employee_id = %s', (employee_id,))
        else:
            cursor.execute('DELETE FROM employees WHERE employee_id = ?', (employee_id,))
        
        conn.commit()
        conn.close()
        
        # Limpiar cache
        system.employees_cache = {}
        system.cache_timestamp = 0
        
        return jsonify({'success': True, 'message': f"Empleado {employee[0]} eliminado exitosamente"})
        
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'message': f"Error: {str(e)}"})

@app.route('/api/test_connection', methods=['POST'])
def api_test_connection():
    connected = system.test_connection()
    message = "✅ Dispositivo conectado" if connected else "❌ Dispositivo no disponible"
    return jsonify({'connected': connected, 'message': message})

@app.route('/api/records')
def api_records():
    date = request.args.get('date')
    if not date:
        return jsonify([])
    
    conn = system.get_connection()
    cursor = conn.cursor()
    
    try:
        if system.db_type == 'postgresql':
            cursor.execute('''
                SELECT e.name, ar.event_type, ar.timestamp, ar.verify_method, e.department
                FROM attendance_records ar
                JOIN employees e ON ar.employee_id = e.employee_id
                WHERE DATE(ar.timestamp) = %s
                ORDER BY ar.timestamp DESC
            ''', (date,))
        else:
            cursor.execute('''
                SELECT e.name, ar.event_type, ar.timestamp, ar.verify_method, e.department
                FROM attendance_records ar
                JOIN employees e ON ar.employee_id = e.employee_id
                WHERE date(ar.timestamp) = ?
                ORDER BY ar.timestamp DESC
            ''', (date,))
        
        records = cursor.fetchall()
        conn.close()
        
        return jsonify([{
            'name': record[0],
            'event_type': record[1],
            'timestamp': record[2],
            'verify_method': record[3],
            'department': record[4] or 'General'
        } for record in records])
        
    except Exception as e:
        conn.close()
        return jsonify([])

@app.route('/api/reports/daily')
def api_daily_report():
    date = request.args.get('date')
    if not date:
        return jsonify({'error': 'Date required'})
    
    conn = system.get_connection()
    cursor = conn.cursor()
    
    try:
        if system.db_type == 'postgresql':
            cursor.execute("SELECT COUNT(*) FROM attendance_records WHERE DATE(timestamp) = %s", (date,))
            total_records = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(DISTINCT employee_id) FROM attendance_records WHERE DATE(timestamp) = %s", (date,))
            unique_employees = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM attendance_records WHERE DATE(timestamp) = %s AND event_type = 'entrada'", (date,))
            entries = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM attendance_records WHERE DATE(timestamp) = %s AND event_type = 'salida'", (date,))
            exits = cursor.fetchone()[0]
            
            cursor.execute('''
                SELECT e.name, ar.event_type, ar.timestamp
                FROM attendance_records ar
                JOIN employees e ON ar.employee_id = e.employee_id
                WHERE DATE(ar.timestamp) = %s
                ORDER BY ar.timestamp
            ''', (date,))
            records = cursor.fetchall()
        else:
            cursor.execute("SELECT COUNT(*) FROM attendance_records WHERE date(timestamp) = ?", (date,))
            total_records = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(DISTINCT employee_id) FROM attendance_records WHERE date(timestamp) = ?", (date,))
            unique_employees = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM attendance_records WHERE date(timestamp) = ? AND event_type = 'entrada'", (date,))
            entries = cursor.fetchone()[0]
            
            cursor.execute("SELECT COUNT(*) FROM attendance_records WHERE date(timestamp) = ? AND event_type = 'salida'", (date,))
            exits = cursor.fetchone()[0]
            
            cursor.execute('''
                SELECT e.name, ar.event_type, ar.timestamp
                FROM attendance_records ar
                JOIN employees e ON ar.employee_id = e.employee_id
                WHERE date(ar.timestamp) = ?
                ORDER BY ar.timestamp
            ''', (date,))
            records = cursor.fetchall()
        
        conn.close()
        
        return jsonify({
            'total_records': total_records,
            'unique_employees': unique_employees,
            'entries': entries,
            'exits': exits,
            'records': [{
                'name': record[0],
                'event_type': record[1],
                'timestamp': record[2]
            } for record in records]
        })
        
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)})

@app.route('/api/reports/weekly')
def api_weekly_report():
    week = request.args.get('week')
    if not week:
        return jsonify({'error': 'Week required'})
    
    # Calcular fechas de la semana
    from datetime import datetime, timedelta
    year, week_num = map(int, week.split('-W'))
    start_date = datetime.strptime(f'{year}-W{week_num}-1', '%Y-W%W-%w')
    end_date = start_date + timedelta(days=6)
    
    conn = system.get_connection()
    cursor = conn.cursor()
    
    try:
        if system.db_type == 'postgresql':
            cursor.execute('''
                SELECT COUNT(*) FROM attendance_records 
                WHERE DATE(timestamp) BETWEEN %s AND %s
            ''', (start_date.date(), end_date.date()))
            total_records = cursor.fetchone()[0]
            
            cursor.execute('''
                SELECT COUNT(DISTINCT employee_id) FROM attendance_records 
                WHERE DATE(timestamp) BETWEEN %s AND %s
            ''', (start_date.date(), end_date.date()))
            active_employees = cursor.fetchone()[0]
        else:
            cursor.execute('''
                SELECT COUNT(*) FROM attendance_records 
                WHERE date(timestamp) BETWEEN ? AND ?
            ''', (start_date.date(), end_date.date()))
            total_records = cursor.fetchone()[0]
            
            cursor.execute('''
                SELECT COUNT(DISTINCT employee_id) FROM attendance_records 
                WHERE date(timestamp) BETWEEN ? AND ?
            ''', (start_date.date(), end_date.date()))
            active_employees = cursor.fetchone()[0]
        
        conn.close()
        
        return jsonify({
            'total_records': total_records,
            'active_employees': active_employees,
            'start_date': start_date.date().isoformat(),
            'end_date': end_date.date().isoformat()
        })
        
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)})

@app.route('/api/reports/attendance')
def api_attendance_report():
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    employee_id = request.args.get('employee_id')
    department = request.args.get('department')
    
    if not start_date or not end_date:
        return jsonify({'error': 'Fechas requeridas'})
    
    try:
        report_data = system.generate_attendance_report(start_date, end_date, employee_id, department)
        return jsonify(report_data)
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/breaks/status')
def api_break_status():
    """Obtener estado actual de breaks y almuerzos"""
    try:
        conn = system.get_connection()
        cursor = conn.cursor()
        
        today = datetime.now().strftime('%Y-%m-%d')
        
        # Empleados actualmente en break
        if system.db_type == 'postgresql':
            cursor.execute('''
                WITH last_break_events AS (
                    SELECT DISTINCT ON (ar.employee_id) 
                           ar.employee_id, e.name, e.department, ar.event_type, ar.timestamp, ar.break_type
                    FROM attendance_records ar
                    JOIN employees e ON ar.employee_id = e.employee_id
                    WHERE DATE(ar.timestamp) = %s 
                          AND ar.is_break_record = true
                          AND e.active = true
                    ORDER BY ar.employee_id, ar.timestamp DESC
                )
                SELECT employee_id, name, department, event_type, timestamp, break_type
                FROM last_break_events
                WHERE event_type IN ('break_salida', 'almuerzo_salida')
            ''', (today,))
        else:
            cursor.execute('''
                SELECT ar.employee_id, e.name, e.department, ar.event_type, ar.timestamp, ar.break_type
                FROM attendance_records ar
                JOIN employees e ON ar.employee_id = e.employee_id
                WHERE date(ar.timestamp) = ? 
                      AND ar.is_break_record = 1
                      AND e.active = 1
                      AND ar.event_type IN ('break_salida', 'almuerzo_salida')
                      AND ar.timestamp = (
                          SELECT MAX(timestamp) FROM attendance_records ar2
                          WHERE ar2.employee_id = ar.employee_id 
                                AND date(ar2.timestamp) = ?
                                AND ar2.is_break_record = 1
                      )
            ''', (today, today))
        
        current_breaks = cursor.fetchall()
        
        # Procesar empleados en break/almuerzo
        on_break = []
        on_lunch = []
        
        for record in current_breaks:
            emp_id, name, department, event_type, timestamp, break_type = record
            
            # Calcular duración
            start_time = datetime.strptime(str(timestamp)[11:19], '%H:%M:%S').time()
            current_time = datetime.now().time()
            
            start_dt = datetime.combine(datetime.today(), start_time)
            current_dt = datetime.combine(datetime.today(), current_time)
            duration = int((current_dt - start_dt).total_seconds() / 60)
            
            employee_data = {
                'employee_id': emp_id,
                'name': name,
                'department': department,
                'start_time': str(start_time)[:5],
                'duration': duration
            }
            
            if event_type == 'break_salida':
                on_break.append(employee_data)
            elif event_type == 'almuerzo_salida':
                on_lunch.append(employee_data)
        
        # Contar breaks completados
        if system.db_type == 'postgresql':
            cursor.execute('''
                SELECT COUNT(*) FROM attendance_records ar
                JOIN employees e ON ar.employee_id = e.employee_id
                WHERE DATE(ar.timestamp) = %s 
                      AND ar.is_break_record = true
                      AND ar.event_type = 'break_entrada'
                      AND e.active = true
            ''', (today,))
        else:
            cursor.execute('''
                SELECT COUNT(*) FROM attendance_records ar
                JOIN employees e ON ar.employee_id = e.employee_id
                WHERE date(ar.timestamp) = ? 
                      AND ar.is_break_record = 1
                      AND ar.event_type = 'break_entrada'
                      AND e.active = 1
            ''', (today,))
        
        breaks_completed = cursor.fetchone()[0]
        
        # Contar almuerzos completados
        if system.db_type == 'postgresql':
            cursor.execute('''
                SELECT COUNT(*) FROM attendance_records ar
                JOIN employees e ON ar.employee_id = e.employee_id
                WHERE DATE(ar.timestamp) = %s 
                      AND ar.is_break_record = true
                      AND ar.event_type = 'almuerzo_entrada'
                      AND e.active = true
            ''', (today,))
        else:
            cursor.execute('''
                SELECT COUNT(*) FROM attendance_records ar
                JOIN employees e ON ar.employee_id = e.employee_id
                WHERE date(ar.timestamp) = ? 
                      AND ar.is_break_record = 1
                      AND ar.event_type = 'almuerzo_entrada'
                      AND e.active = 1
            ''', (today,))
        
        lunch_completed = cursor.fetchone()[0]
        
        # Calcular pendientes
        if system.db_type == 'postgresql':
            cursor.execute("SELECT COUNT(*) FROM employees WHERE active = true AND department IN ('Reacondicionamiento', 'Logistica', 'Administracion')")
        else:
            cursor.execute("SELECT COUNT(*) FROM employees WHERE active = 1 AND department IN ('Reacondicionamiento', 'Logistica', 'Administracion')")
        
        admin_employees = cursor.fetchone()[0]
        
        if system.db_type == 'postgresql':
            cursor.execute("SELECT COUNT(*) FROM employees WHERE active = true AND department = 'Operativos'")
        else:
            cursor.execute("SELECT COUNT(*) FROM employees WHERE active = 1 AND department = 'Operativos'")
        
        operativo_employees = cursor.fetchone()[0]
        
        total_employees = admin_employees + operativo_employees
        breaks_pending = max(0, total_employees - breaks_completed)
        lunch_pending = max(0, admin_employees - lunch_completed)
        
        conn.close()
        
        return jsonify({
            'on_break': on_break,
            'on_lunch': on_lunch,
            'breaks_completed': breaks_completed,
            'breaks_pending': breaks_pending,
            'lunch_completed': lunch_completed,
            'lunch_pending': lunch_pending
        })
        
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/alerts/late')
def api_late_alerts():
    """Obtener alertas de llegadas tardías del día con horarios por departamento"""
    try:
        conn = system.get_connection()
        cursor = conn.cursor()
        
        today = datetime.now().strftime('%Y-%m-%d')
        
        if system.db_type == 'postgresql':
            cursor.execute('''
                SELECT e.employee_id, e.name, e.department, e.schedule,
                       MIN(ar.timestamp) as first_entry
                FROM employees e
                JOIN attendance_records ar ON e.employee_id = ar.employee_id
                WHERE DATE(ar.timestamp) = %s AND ar.event_type = 'entrada' AND e.active = true
                GROUP BY e.employee_id, e.name, e.department, e.schedule
            ''', (today,))
        else:
            cursor.execute('''
                SELECT e.employee_id, e.name, e.department, e.schedule,
                       MIN(ar.timestamp) as first_entry
                FROM employees e
                JOIN attendance_records ar ON e.employee_id = ar.employee_id
                WHERE date(ar.timestamp) = ? AND ar.event_type = 'entrada' AND e.active = 1
                GROUP BY e.employee_id, e.name, e.department, e.schedule
            ''', (today,))
        
        records = cursor.fetchall()
        conn.close()
        
        late_alerts = []
        
        for record in records:
            emp_id, name, dept, schedule, first_entry = record
            
            # Obtener horario esperado por departamento
            entry_time = datetime.strptime(str(first_entry)[:19], '%Y-%m-%d %H:%M:%S')
            day_of_week = entry_time.weekday()
            
            expected_hours = system.get_expected_hours_by_department(dept, day_of_week)
            if not expected_hours:
                continue
                
            # Extraer hora de entrada
            entry_time_only = entry_time.time()
            expected_start = datetime.strptime(expected_hours[0], '%H:%M').time()
            
            if entry_time_only > expected_start:
                # Calcular minutos de tardanza
                entry_dt = datetime.combine(datetime.today(), entry_time_only)
                expected_dt = datetime.combine(datetime.today(), expected_start)
                late_minutes = int((entry_dt - expected_dt).total_seconds() / 60)
                
                late_alerts.append({
                    'employee_id': emp_id,
                    'name': name,
                    'department': dept,
                    'expected_time': expected_hours[0],
                    'actual_time': str(entry_time_only)[:5],
                    'late_minutes': late_minutes,
                    'timestamp': first_entry
                })
        
        return jsonify(late_alerts)
        
    except Exception as e:
        return jsonify({'error': str(e)})
@app.route('/api/export/excel')
def api_export_excel():
    if not EXCEL_AVAILABLE:
        return jsonify({'error': 'Exportación Excel no disponible. Instala: pip install openpyxl'})
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    employee_id = request.args.get('employee_id')
    department = request.args.get('department')
    
    if not start_date or not end_date:
        return jsonify({'error': 'Fechas requeridas'})
    
    try:
        report_data = system.generate_attendance_report(start_date, end_date, employee_id, department)
        
        print(f"Debug: report_data keys: {list(report_data.keys()) if report_data else 'Empty'}")
        print(f"Debug: start_date={start_date}, end_date={end_date}, employee_id={employee_id}, department={department}")
        
        if not report_data:
            return jsonify({'error': 'No se encontraron empleados o registros para el rango de fechas seleccionado'})
        
        filename = f"reporte_asistencia_{start_date}_{end_date}.xlsx"
        filepath = os.path.join('exports', filename)
        
        # Crear directorio si no existe
        os.makedirs('exports', exist_ok=True)
        
        system.export_to_excel(report_data, filepath)
        
        return send_file(filepath, as_attachment=True, download_name=filename)
        
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/export/pdf')
def api_export_pdf():
    if not PDF_AVAILABLE:
        return jsonify({'error': 'Exportación PDF no disponible. Instala: pip install reportlab'})
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    employee_id = request.args.get('employee_id')
    department = request.args.get('department')
    
    if not start_date or not end_date:
        return jsonify({'error': 'Fechas requeridas'})
    
    try:
        report_data = system.generate_attendance_report(start_date, end_date, employee_id, department)
        
        if not report_data:
            return jsonify({'error': 'No hay datos para exportar'})
        
        filename = f"reporte_asistencia_{start_date}_{end_date}.pdf"
        filepath = os.path.join('exports', filename)
        
        # Crear directorio si no existe
        os.makedirs('exports', exist_ok=True)
        
        system.export_to_pdf(report_data, filepath)
        
        return send_file(filepath, as_attachment=True, download_name=filename)
        
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/schedules/bulk', methods=['POST'])
def api_bulk_assign_schedule():
    """Asignar horarios en lote a múltiples empleados"""
    data = request.json
    conn = system.get_connection()
    cursor = conn.cursor()
    
    try:
        employee_ids = data.get('employee_ids', [])
        shift_type = data.get('shift_type')
        week_start = data.get('week_start')  # YYYY-MM-DD
        
        if not employee_ids or not shift_type:
            return jsonify({'success': False, 'message': 'Empleados y turno requeridos'})
        
        # Verificar si algún empleado ya tiene turno asignado para esa semana
        from datetime import datetime, timedelta
        week_start_date = datetime.strptime(week_start, '%Y-%m-%d').date()
        week_end_date = week_start_date + timedelta(days=6)
        
        # Verificar conflictos
        conflicts = []
        for emp_id in employee_ids:
            if system.db_type == 'postgresql':
                cursor.execute('''
                    SELECT e.name, wsa.shift_type FROM weekly_shift_assignments wsa
                    JOIN employees e ON wsa.employee_id = e.employee_id
                    WHERE wsa.employee_id = %s AND wsa.week_start = %s
                ''', (emp_id, week_start_date))
            else:
                cursor.execute('''
                    SELECT e.name, wsa.shift_type FROM weekly_shift_assignments wsa
                    JOIN employees e ON wsa.employee_id = e.employee_id
                    WHERE wsa.employee_id = ? AND wsa.week_start = ?
                ''', (emp_id, week_start_date))
            
            existing = cursor.fetchone()
            if existing:
                conflicts.append(f"{existing[0]} ya tiene turno {existing[1]} asignado")
        
        if conflicts:
            return jsonify({
                'success': False, 
                'message': f'Conflictos encontrados: {"; ".join(conflicts)}'
            })
        
        # Configurar horarios
        shift_configs = {
            'mañana': {'start_time': '06:00:00', 'end_time': '14:00:00'},
            'tarde': {'start_time': '14:00:00', 'end_time': '21:00:00'},
            'noche': {'start_time': '22:00:00', 'end_time': '06:00:00'}
        }
        
        if shift_type not in shift_configs:
            return jsonify({'success': False, 'message': 'Turno inválido'})
        
        config = shift_configs[shift_type]
        
        # Insertar asignaciones semanales
        for emp_id in employee_ids:
            if system.db_type == 'postgresql':
                cursor.execute('''
                    INSERT INTO weekly_shift_assignments 
                    (employee_id, week_start, week_end, shift_type, start_time, end_time)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    ON CONFLICT (employee_id, week_start) 
                    DO UPDATE SET 
                        shift_type = EXCLUDED.shift_type,
                        start_time = EXCLUDED.start_time,
                        end_time = EXCLUDED.end_time
                ''', (emp_id, week_start_date, week_end_date, shift_type, 
                      config['start_time'], config['end_time']))
            else:
                cursor.execute('''
                    INSERT OR REPLACE INTO weekly_shift_assignments 
                    (employee_id, week_start, week_end, shift_type, start_time, end_time)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (emp_id, week_start_date, week_end_date, shift_type, 
                      config['start_time'], config['end_time']))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True, 
            'message': f'Turno {shift_type} asignado a {len(employee_ids)} empleados para la semana {week_start}'
        })
        
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/employees/technicians')
def api_get_technicians():
    """Obtener técnicos del departamento Desarme con estado de asignación"""
    conn = system.get_connection()
    cursor = conn.cursor()
    
    try:
        week_start = request.args.get('week_start')
        
        if system.db_type == 'postgresql':
            if week_start:
                cursor.execute('''
                    SELECT e.employee_id, e.name, e.department, wsa.shift_type
                    FROM employees e
                    LEFT JOIN weekly_shift_assignments wsa ON e.employee_id = wsa.employee_id AND wsa.week_start = %s
                    WHERE e.active = true AND e.department = 'Operativos'
                    ORDER BY e.name
                ''', (week_start,))
            else:
                cursor.execute('''
                    SELECT employee_id, name, department, NULL as shift_type
                    FROM employees 
                    WHERE active = true AND department = 'Operativos'
                    ORDER BY name
                ''')
        else:
            if week_start:
                cursor.execute('''
                    SELECT e.employee_id, e.name, e.department, wsa.shift_type
                    FROM employees e
                    LEFT JOIN weekly_shift_assignments wsa ON e.employee_id = wsa.employee_id AND wsa.week_start = ?
                    WHERE e.active = 1 AND e.department = 'Operativos'
                    ORDER BY e.name
                ''', (week_start,))
            else:
                cursor.execute('''
                    SELECT employee_id, name, department, NULL as shift_type
                    FROM employees 
                    WHERE active = 1 AND department = 'Operativos'
                    ORDER BY name
                ''')
        
        technicians = cursor.fetchall()
        conn.close()
        
        return jsonify([{
            'employee_id': tech[0],
            'name': tech[1],
            'department': tech[2],
            'assigned_shift': tech[3] if len(tech) > 3 else None
        } for tech in technicians])
        
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)})

@app.route('/api/schedules/weekly-report')
def api_weekly_schedule_report():
    """Obtener reporte semanal de turnos"""
    week_start = request.args.get('week_start')
    if not week_start:
        return jsonify({'error': 'Fecha de inicio de semana requerida'})
    
    try:
        conn = system.get_connection()
        cursor = conn.cursor()
        
        from datetime import datetime, timedelta
        week_start_date = datetime.strptime(week_start, '%Y-%m-%d').date()
        week_end_date = week_start_date + timedelta(days=6)
        
        if system.db_type == 'postgresql':
            cursor.execute('''
                SELECT e.employee_id, e.name, e.department, wsa.shift_type, wsa.start_time, wsa.end_time
                FROM employees e
                JOIN weekly_shift_assignments wsa ON e.employee_id = wsa.employee_id
                WHERE e.active = true AND wsa.week_start = %s
                ORDER BY wsa.shift_type, e.name
            ''', (week_start_date,))
        else:
            cursor.execute('''
                SELECT e.employee_id, e.name, e.department, wsa.shift_type, wsa.start_time, wsa.end_time
                FROM employees e
                JOIN weekly_shift_assignments wsa ON e.employee_id = wsa.employee_id
                WHERE e.active = 1 AND wsa.week_start = ?
                ORDER BY wsa.shift_type, e.name
            ''', (week_start_date,))
        
        schedules = cursor.fetchall()
        conn.close()
        
        # Agrupar por turno
        report = {
            'week_start': week_start,
            'week_end': week_end_date.strftime('%Y-%m-%d'),
            'shifts': {
                'mañana': [],
                'tarde': [],
                'noche': []
            }
        }
        
        for schedule in schedules:
            emp_id, name, dept, shift, start_time, end_time = schedule
            if shift in report['shifts']:
                report['shifts'][shift].append({
                    'employee_id': emp_id,
                    'name': name,
                    'department': dept,
                    'start_time': str(start_time),
                    'end_time': str(end_time)
                })
        
        return jsonify(report)
        
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/schedules/export-pdf')
def api_export_schedule_pdf():
    """Exportar horarios semanales a PDF"""
    if not PDF_AVAILABLE:
        return jsonify({'error': 'Exportación PDF no disponible'})
    
    week_start = request.args.get('week_start')
    if not week_start:
        return jsonify({'error': 'Fecha requerida'})
    
    try:
        # Obtener datos del reporte
        conn = system.get_connection()
        cursor = conn.cursor()
        
        from datetime import datetime, timedelta
        week_start_date = datetime.strptime(week_start, '%Y-%m-%d').date()
        week_end_date = week_start_date + timedelta(days=6)
        
        if system.db_type == 'postgresql':
            cursor.execute('''
                SELECT e.employee_id, e.name, e.department, wsa.shift_type, wsa.start_time, wsa.end_time
                FROM employees e
                JOIN weekly_shift_assignments wsa ON e.employee_id = wsa.employee_id
                WHERE e.active = true AND wsa.week_start = %s
                ORDER BY wsa.shift_type, e.name
            ''', (week_start_date,))
        else:
            cursor.execute('''
                SELECT e.employee_id, e.name, e.department, wsa.shift_type, wsa.start_time, wsa.end_time
                FROM employees e
                JOIN weekly_shift_assignments wsa ON e.employee_id = wsa.employee_id
                WHERE e.active = 1 AND wsa.week_start = ?
                ORDER BY wsa.shift_type, e.name
            ''', (week_start_date,))
        
        schedules = cursor.fetchall()
        conn.close()
        
        # Crear PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Título
        title = Paragraph(f"<b>HORARIOS SEMANALES - PCSHEK</b><br/>Semana del {week_start_date.strftime('%d/%m/%Y')} al {week_end_date.strftime('%d/%m/%Y')}", styles['Title'])
        story.append(title)
        story.append(Spacer(1, 20))
        
        # Agrupar por turno
        shifts = {'mañana': [], 'tarde': [], 'noche': []}
        for schedule in schedules:
            emp_id, name, dept, shift, start_time, end_time = schedule
            if shift in shifts:
                shifts[shift].append([name, dept, f"{start_time} - {end_time}"])
        
        # Crear tablas por turno
        shift_names = {'mañana': 'TURNO MAÑANA (06:00 - 14:00)', 'tarde': 'TURNO TARDE (14:00 - 21:00)', 'noche': 'TURNO NOCHE (22:00 - 06:00)'}
        
        for shift_key, shift_name in shift_names.items():
            if shifts[shift_key]:
                # Subtítulo
                subtitle = Paragraph(f"<b>{shift_name}</b>", styles['Heading2'])
                story.append(subtitle)
                story.append(Spacer(1, 10))
                
                # Tabla
                data = [['Empleado', 'Departamento', 'Horario']]
                data.extend(shifts[shift_key])
                
                table = Table(data, colWidths=[200, 150, 100])
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 10),
                ]))
                
                story.append(table)
                story.append(Spacer(1, 20))
        
        doc.build(story)
        
        filename = f"horarios_semana_{week_start}.pdf"
        filepath = os.path.join('exports', filename)
        os.makedirs('exports', exist_ok=True)
        
        with open(filepath, 'wb') as f:
            f.write(buffer.getvalue())
        
        return send_file(filepath, as_attachment=True, download_name=filename)
        
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/schedules')
def api_get_schedules():
    conn = system.get_connection()
    cursor = conn.cursor()
    
    try:
        if system.db_type == 'postgresql':
            cursor.execute('''
                SELECT es.employee_id, e.name, es.schedule_type, es.shift_type, 
                       es.start_time, es.end_time, es.days_of_week, es.active_from, es.active_until
                FROM employee_schedules es
                JOIN employees e ON es.employee_id = e.employee_id
                WHERE e.active = true
                ORDER BY e.name
            ''')
        else:
            cursor.execute('''
                SELECT es.employee_id, e.name, es.schedule_type, es.shift_type, 
                       es.start_time, es.end_time, es.days_of_week, es.active_from, es.active_until
                FROM employee_schedules es
                JOIN employees e ON es.employee_id = e.employee_id
                WHERE e.active = 1
                ORDER BY e.name
            ''')
        
        schedules = cursor.fetchall()
        conn.close()
        
        return jsonify([{
            'employee_id': schedule[0],
            'name': schedule[1],
            'schedule_type': schedule[2],
            'shift_type': schedule[3],
            'start_time': str(schedule[4]),
            'end_time': str(schedule[5]),
            'days_of_week': schedule[6],
            'active_from': str(schedule[7]) if schedule[7] else None,
            'active_until': str(schedule[8]) if schedule[8] else None
        } for schedule in schedules])
        
    except Exception as e:
        conn.close()
        return jsonify({'error': str(e)})

@app.route('/api/reports/monthly')
def api_monthly_report():
    """Obtener reporte mensual para nómina"""
    month = request.args.get('month')  # YYYY-MM
    employee_id = request.args.get('employee_id')
    
    if not month:
        return jsonify({'error': 'Mes requerido (formato YYYY-MM)'})
    
    try:
        conn = system.get_connection()
        cursor = conn.cursor()
        
        # Construir consulta
        if system.db_type == 'postgresql':
            base_query = '''
                SELECT e.employee_id, e.name, e.department, e.schedule,
                       ds.date, ds.first_entry, ds.last_exit, ds.total_hours, 
                       ds.worked_day, ds.is_weekend, ds.is_holiday
                FROM employees e
                LEFT JOIN daily_summaries ds ON e.employee_id = ds.employee_id
                WHERE e.active = true AND DATE_TRUNC('month', ds.date) = %s::date
            '''
            params = [month + '-01']
        else:
            base_query = '''
                SELECT e.employee_id, e.name, e.department, e.schedule,
                       ds.date, ds.first_entry, ds.last_exit, ds.total_hours, 
                       ds.worked_day, ds.is_weekend, ds.is_holiday
                FROM employees e
                LEFT JOIN daily_summaries ds ON e.employee_id = ds.employee_id
                WHERE e.active = 1 AND strftime('%Y-%m', ds.date) = ?
            '''
            params = [month]
        
        if employee_id:
            base_query += ' AND e.employee_id = {}'
            base_query = base_query.format('%s' if system.db_type == 'postgresql' else '?')
            params.append(employee_id)
        
        base_query += ' ORDER BY e.name, ds.date'
        
        cursor.execute(base_query, params)
        records = cursor.fetchall()
        conn.close()
        
        # Procesar datos por empleado
        monthly_data = {}
        
        for record in records:
            emp_id, name, dept, schedule, date, entry, exit, hours, worked, weekend, holiday = record
            
            if emp_id not in monthly_data:
                monthly_data[emp_id] = {
                    'employee_id': emp_id,
                    'name': name,
                    'department': dept,
                    'schedule': schedule,
                    'total_days_worked': 0,
                    'total_hours': 0,
                    'weekend_days': 0,
                    'holiday_days': 0,
                    'days': []
                }
            
            if date:
                monthly_data[emp_id]['days'].append({
                    'date': str(date),
                    'first_entry': str(entry) if entry else None,
                    'last_exit': str(exit) if exit else None,
                    'total_hours': float(hours) if hours else 0,
                    'worked_day': bool(worked),
                    'is_weekend': bool(weekend),
                    'is_holiday': bool(holiday)
                })
                
                if worked:
                    monthly_data[emp_id]['total_days_worked'] += 1
                    monthly_data[emp_id]['total_hours'] += float(hours) if hours else 0
                
                if weekend:
                    monthly_data[emp_id]['weekend_days'] += 1
                
                if holiday:
                    monthly_data[emp_id]['holiday_days'] += 1
        
        return jsonify(monthly_data)
        
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/reports/monthly-summary')
def api_monthly_summary():
    """Reporte mensual resumido: Empleado, Departamento, Días Presente, Días Ausente, Horas Totales"""
    month = request.args.get('month')  # YYYY-MM
    
    if not month:
        return jsonify({'error': 'Mes requerido (formato YYYY-MM)'})
    
    try:
        conn = system.get_connection()
        cursor = conn.cursor()
        
        # Obtener todos los empleados activos
        if system.db_type == 'postgresql':
            cursor.execute('SELECT employee_id, name, department FROM employees WHERE active = true ORDER BY name')
        else:
            cursor.execute('SELECT employee_id, name, department FROM employees WHERE active = 1 ORDER BY name')
        
        employees = cursor.fetchall()
        
        # Calcular días laborables del mes (excluyendo fines de semana)
        year, month_num = map(int, month.split('-'))
        import calendar
        
        # Contar días laborables (L-V)
        work_days = 0
        for day in range(1, calendar.monthrange(year, month_num)[1] + 1):
            weekday = calendar.weekday(year, month_num, day)
            if weekday < 5:  # L-V (0-4)
                work_days += 1
        
        summary_data = []
        
        for emp_id, name, department in employees:
            # Obtener resumen del empleado para el mes
            if system.db_type == 'postgresql':
                cursor.execute('''
                    SELECT COUNT(*) as days_present, COALESCE(SUM(total_hours), 0) as total_hours
                    FROM daily_summaries 
                    WHERE employee_id = %s AND DATE_TRUNC('month', date) = %s::date AND worked_day = true
                ''', (emp_id, month + '-01'))
            else:
                cursor.execute('''
                    SELECT COUNT(*) as days_present, COALESCE(SUM(total_hours), 0) as total_hours
                    FROM daily_summaries 
                    WHERE employee_id = ? AND strftime('%Y-%m', date) = ? AND worked_day = 1
                ''', (emp_id, month))
            
            result = cursor.fetchone()
            days_present = result[0] if result else 0
            total_hours = round(float(result[1]) if result and result[1] else 0, 2)
            days_absent = work_days - days_present
            
            summary_data.append({
                'employee_id': emp_id,
                'name': name,
                'department': department,
                'days_present': days_present,
                'days_absent': max(0, days_absent),
                'total_hours': total_hours
            })
        
        conn.close()
        
        return jsonify({
            'month': month,
            'work_days': work_days,
            'employees': summary_data,
            'totals': {
                'total_employees': len(summary_data),
                'total_hours': sum(emp['total_hours'] for emp in summary_data),
                'avg_days_present': round(sum(emp['days_present'] for emp in summary_data) / len(summary_data), 1) if summary_data else 0
            }
        })
        
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/schedules/<employee_id>', methods=['DELETE'])
def api_delete_schedule(employee_id):
    conn = system.get_connection()
    cursor = conn.cursor()
    
    try:
        if system.db_type == 'postgresql':
            cursor.execute('DELETE FROM employee_schedules WHERE employee_id = %s', (employee_id,))
        else:
            cursor.execute('DELETE FROM employee_schedules WHERE employee_id = ?', (employee_id,))
        
        if cursor.rowcount > 0:
            conn.commit()
            conn.close()
            return jsonify({'success': True, 'message': 'Horario eliminado exitosamente'})
        else:
            conn.close()
            return jsonify({'success': False, 'message': 'Horario no encontrado'})
            
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

if __name__ == '__main__':
    print("SISTEMA DE ASISTENCIA OPTIMIZADO")
    print("=" * 50)
    print(f"Dispositivo: {system.device_ip}")
    print(f"Base de datos: {system.db_type.upper()}")
    print(f"Dashboard: http://localhost:5000")
    print("=" * 50)
    
    # Probar conexión inicial
    if system.test_connection():
        print("Dispositivo conectado - Monitoreo activo")
        system.start_monitoring()
    else:
        print("Dispositivo no disponible - iniciando sin monitoreo")
    
    # Configuración para producción
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV') != 'production'
    
    socketio.run(app, debug=debug, host='0.0.0.0', port=port, allow_unsafe_werkzeug=True)